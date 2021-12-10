#!/usr/bin/env python3

# build an Excel file with 3 sheets

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# init workbook
wb = Workbook()

# default first sheet
ws1 = wb.active
ws1.title = 'Sheet 1 name'
# fill it
for col in 'CDEFGH':
    ws1.column_dimensions[col].width = 5
    for row in range(3, 10):
        cell_ref = f'{col}{row}'
        c = ws1[cell_ref]
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(top=Side(border_style='double', color='ff0000'))
        c.value = f'{c.coordinate}'
        # odd or even lines formatting
        if row % 2:
            c.fill = PatternFill('solid', fgColor='dddddd')
            c.font = Font(b=False, color='ff0000')
        else:
            c.fill = PatternFill('solid', fgColor='eeeeee')
            c.font = Font(b=False, color='0000ed')

# create sheet #2
ws2 = wb.create_sheet(title='Pi')
ws2['C3'] = 3.14

# create sheet #3
ws3 = wb.create_sheet(title='Data')
for row in range(10, 20):
    for col in range(5, 10):
        ws3.cell(column=col, row=row, value=f'{get_column_letter(col)}')

# save all to xlsx file
wb.save(filename='example.xlsx')
