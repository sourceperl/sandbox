#!/usr/bin/env python3

""" Example of time series data on a sheet with a line chart. """

from datetime import datetime
import os
import sys
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference


# rows samples
rows = [
    ['Date and time', 'Field 1', 'Field 2', 'Field 3'],
    ['2019-01-04T16:41:24', 40, 30, 25],
    ['2019-01-05T16:41:24', 30, 30, 25],
    ['2019-01-07T17:41:24', 40, 25, 30],
]

# convert iso str to datetime
for row in rows[1:]:
    row[0] = datetime.fromisoformat(row[0])

# init a workbook with main sheet as ws objet
wb = Workbook()
ws = wb.active

# add rows to main sheet
for row in rows:
    ws.append(row)

# increase the size of the "Date and time" column
ws.column_dimensions['A'].width = 22.0

# build a line chart
mychart = LineChart()
mychart.style = 12
# add axis names
mychart.x_axis.title = 'Time'
mychart.y_axis.title = 'Field value'
# add "field xx" values as chart data (keep this before set_categories)
data = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
mychart.add_data(data, from_rows=False, titles_from_data=True)
# add "date and time" as chart categories (on x-axis)
cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)
mychart.set_categories(cats)
# add this chart to main sheet
ws.add_chart(mychart, 'A8')

# save xlsx
xlsx_dt_str = datetime.strftime(datetime.now(), '%Y%m%d_%H%M%S')
xlsx_name = f'example-{xlsx_dt_str}.xlsx'
wb.save(xlsx_name)

# open xlsx with viewer app
if sys.platform == 'win32':
    os.system(f'start excel {xlsx_name}')
elif sys.platform == 'linux':
    os.system(f'libreoffice {xlsx_name}')
