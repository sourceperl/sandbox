#!/usr/bin/env python3

""" Export influxdb data to a zipped Excel workbook. """

from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile
from zipfile import ZipFile, ZIP_DEFLATED
# apt install python3-influxdb
from influxdb import InfluxDBClient
# apt install python3-openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# connect to influxdb DB
client = InfluxDBClient(host='192.168.0.24', port=8086, database='mydb')

# do influx request
tags_values_d = {}
req = 'SELECT * FROM "tags" ORDER BY time DESC LIMIT 20000'
for point in client.query(req).get_points():
    # format result as a tags dict
    try:
        tags_values_d[point['tag']].append((point['time'], point['value']))
    except KeyError:
        tags_values_d[point['tag']] = [(point['time'], point['value'])]

# store result in a xlsx workbook
wb = Workbook()

# format cover sheet
first_sheet = wb.active
first_sheet.title = 'Main'
first_sheet.row_dimensions[1].height = 40
first_sheet.column_dimensions['A'].width = 30
first_sheet.column_dimensions['B'].width = 25
first_sheet.column_dimensions['C'].width = 25
first_sheet.merge_cells('A1:C1')
first_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
first_sheet['A1'].font = Font(bold=True, underline='single', size=14)
first_sheet['A1'] = 'Test de l\'analyseur Hemera G800 - site de Y'
first_sheet['A5'].alignment = Alignment(horizontal='right')
first_sheet['A5'] = 'Date d\'export des données:'
first_sheet['B5'].alignment = Alignment(horizontal='center')
first_sheet['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# populate right sheets with tags value
for tag_name, tag_history in sorted(tags_values_d.items()):
    # init sheet
    sheet = wb.create_sheet(title=tag_name)
    sheet.row_dimensions[1].height = 20
    sheet.column_dimensions['A'].width = 30
    sheet['A1'].font = Font(bold=True)
    sheet['A1'] = 'Date et heure (UTC)'
    sheet.column_dimensions['B'].width = 20
    sheet['B1'].font = Font(bold=True)
    sheet['B1'] = 'Valeur'
    # add history values to sheet
    for idx, (date_str, value) in enumerate(tag_history):
        sheet.cell(row=idx+2, column=1, value=f'{date_str}')
        sheet.cell(row=idx+2, column=2, value=f'{value:.2f}')
    # apply alignment to all cells in this sheet
    for column in sheet.columns:
        for cell in column:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# save all to xlsx file
with NamedTemporaryFile() as tmp:
    wb.save(tmp.name)
    tmp.seek(0)
    xlsx_bytes = tmp.read()

# create a zip file in a BytesIO stream an add xlsx into
io_zipfile = BytesIO()
with ZipFile(io_zipfile, 'a', compression=ZIP_DEFLATED, compresslevel=9) as f:
    f.writestr('example.xlsx', xlsx_bytes)

# copy BytesIO stream to a file
open('example.zip', 'wb').write(io_zipfile.getvalue())
