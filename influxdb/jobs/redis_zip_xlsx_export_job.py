#!/usr/bin/env python3

""" Export influxdb data to a zipped Excel workbook. """

import argparse
import logging
import time
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile
from zipfile import ZipFile, ZIP_DEFLATED
# apt install python3-influxdb
from influxdb import InfluxDBClient
# apt install python3-openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
# apt install python3-redis
from redis import StrictRedis
# pip3 install schedule
import schedule

# some consts
REDIS_XLSX_KEY = 'xlsx-export:zip'

# some functions
def get_xlsx_content():
    # do influx request
    tags_values_d = {}
    req = 'SELECT * FROM "tags" ORDER BY time DESC LIMIT 20000'
    for point in ifl_cli.query(req, epoch='s').get_points():
        # format result as a tags dict
        try:
            tags_values_d[point['tag']].append((point['time'], point['value']))
        except KeyError:
            tags_values_d[point['tag']] = [(point['time'], point['value'])]

    # store result in a xlsx workbook
    wb = Workbook()

    # format cover sheet
    first_sheet = wb.active
    first_sheet.title = 'Main'
    first_sheet.row_dimensions[1].height = 40
    first_sheet.column_dimensions['A'].width = 40
    first_sheet.column_dimensions['B'].width = 25
    first_sheet.column_dimensions['C'].width = 20
    # line 1
    first_sheet.merge_cells('A1:C1')
    first_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    first_sheet['A1'].font = Font(bold=True, underline='single', size=14)
    first_sheet['A1'] = 'Test de l\'analyseur Hemera G800 - site de Y'
    # line 5
    first_sheet['A5'].alignment = Alignment(horizontal='right')
    first_sheet['A5'] = 'Date et heure d\'export des données:'
    first_sheet['B5'].alignment = Alignment(horizontal='center')
    first_sheet['B5'] = datetime.now()
    # from line 8
    edited_row = 8
    for tag_name, unit in sorted(dict(FLOW='ml/min', CO2='%', O2='%', CH4='%', H2S='ppm', THT='mg/Nm3').items()):
        first_sheet.cell(row=edited_row, column=1).alignment = Alignment(horizontal='right')
        first_sheet.cell(row=edited_row, column=1, value=f'Unité {tag_name}:')
        first_sheet.cell(row=edited_row, column=2).alignment = Alignment(horizontal='center')
        first_sheet.cell(row=edited_row, column=2, value=f'{unit}')
        edited_row += 1

    # populate right sheets with tags value
    for tag_name, tag_history in sorted(tags_values_d.items()):
        # init sheet
        sheet = wb.create_sheet(title=tag_name)
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions['A'].width = 30
        sheet['A1'].font = Font(bold=True)
        sheet['A1'] = 'Date et heure'
        sheet.column_dimensions['B'].width = 20
        sheet['B1'].font = Font(bold=True)
        sheet['B1'] = 'Valeur'
        # add history values to sheet
        for idx, (date_str, value) in enumerate(tag_history):
            sheet.cell(row=idx + 2, column=1, value=datetime.fromtimestamp(date_str))
            sheet.cell(row=idx + 2, column=2, value=round(value, 2))
        # apply alignment to all cells in this sheet
        for column in sheet.columns:
            for cell in column:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # save all to xlsx file
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        xlsx_bytes = tmp.read()

    # return xlsx content as bytes
    return xlsx_bytes


# schedule jobs
def export_xlsx_to_redis_job():
    """Periodic convert of influx data to a zipped workbook store as bytes in a redis key"""
    # log job run
    logging.info('start export_xlsx_to_redis_job')

    # build workbook
    xlsx_content = get_xlsx_content()

    # create a zip file in a BytesIO stream an add xlsx into
    io_zipfile = BytesIO()
    with ZipFile(io_zipfile, 'a', compression=ZIP_DEFLATED, compresslevel=9) as f:
        f.writestr('example.xlsx', xlsx_content)

    # transfer zip file as a redis key
    logging.info(f'transfer the built zip file to the redis key "{REDIS_XLSX_KEY}"')
    red_cli.set(REDIS_XLSX_KEY, io_zipfile.getvalue())


if __name__ == '__main__':
    # parse command line args
    parser = argparse.ArgumentParser()
    parser.add_argument('filename', nargs='?', default='', help='excel filename')
    args = parser.parse_args()
    # logging setup
    logging.basicConfig(format='%(asctime)s %(message)s', level=logging.INFO)
    # init influxdb
    ifl_cli = InfluxDBClient(host='localhost', port=8086, database='mydb')
    # immediate produce of xlsx file or start redis periodic export
    if args.filename:
        logging.info(f'start build of "{args.filename}" workbook')
        open(args.filename, 'wb').write(get_xlsx_content())
        logging.info(f'build done')
    else:
        logging.info('xlsx-build started')
        # init redis DB
        red_cli = StrictRedis()
        # schedule setup
        schedule.every(1).hour.at('00:00').do(export_xlsx_to_redis_job)
        # main loop
        while True:
            schedule.run_pending()
            time.sleep(1.0)
