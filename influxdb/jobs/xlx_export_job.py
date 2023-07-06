#!/usr/bin/env python3

""" Export influxdb data to workbook. """

import argparse
from datetime import datetime, date, timedelta, timezone
import logging
import os
import time
# apt install python3-influxdb
from influxdb import InfluxDBClient
# apt install python3-openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
# pip3 install schedule
import schedule

# some consts
EXPORT_DIR = '/home/export/export_xlsx'


# some functions
def build_daily_xlsx(xlsx_day: date, to_file: str = ''):
    # create xlsx file name and path
    xlsx_file = to_file
    if not xlsx_file:
        xlsx_name = f'g800_{xlsx_day.year:04}{xlsx_day.month:02}{xlsx_day.day:02}.xlsx'
        xlsx_file = os.path.join(EXPORT_DIR, xlsx_name)
    # log start of build
    logging.info(f'build {xlsx_file}')

    # do influx request
    tags_values_d = {}
    # convert day date to string UTC boundary (from_str and to_str)
    from_dt = datetime.combine(xlsx_day, datetime.min.time())
    from_str = from_dt.astimezone(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    to_dt = from_dt + timedelta(days=1)
    to_str = to_dt.astimezone(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    # build influx request
    req = f'SELECT * FROM "tags" WHERE time > \'{from_str}\' AND time <= \'{to_str}\' ORDER BY time DESC LIMIT 10000'
    for point in ifl_cli.query(req, epoch='s').get_points():
        # format result as a tags dict
        try:
            tags_values_d[point['tag']].append((point['time'], point['value']))
        except KeyError:
            tags_values_d[point['tag']] = [(point['time'], point['value'])]

    # store result in a xlsx workbook
    wb = Workbook()

    # format cover sheet
    first_sheet = wb.active
    first_sheet.title = 'Main'
    first_sheet.row_dimensions[1].height = 40
    first_sheet.column_dimensions['A'].width = 40
    first_sheet.column_dimensions['B'].width = 25
    first_sheet.column_dimensions['C'].width = 20
    # line 1
    first_sheet.merge_cells('A1:C1')
    first_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    first_sheet['A1'].font = Font(bold=True, underline='single', size=14)
    first_sheet['A1'] = 'Test de l\'analyseur Hemera G800 - site de Y'
    # line 5
    first_sheet['A5'].alignment = Alignment(horizontal='right')
    first_sheet['A5'] = 'Date et heure d\'export des données:'
    first_sheet['B5'].alignment = Alignment(horizontal='center')
    first_sheet['B5'] = datetime.now()
    # from line 8
    edited_row = 8
    for tag_name, unit in sorted(dict(FLOW='ml/min', CO2='%', O2='%', CH4='%', H2S='ppm', THT='ppm').items()):
        first_sheet.cell(row=edited_row, column=1).alignment = Alignment(horizontal='right')
        first_sheet.cell(row=edited_row, column=1, value=f'Unité {tag_name}:')
        first_sheet.cell(row=edited_row, column=2).alignment = Alignment(horizontal='center')
        first_sheet.cell(row=edited_row, column=2, value=f'{unit}')
        edited_row += 1

    # populate right sheets with tags value
    for tag_name, tag_history in sorted(tags_values_d.items()):
        # THT test
        is_tht = tag_name.startswith('THT')
        # init sheet
        sheet = wb.create_sheet(title=tag_name)
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions['A'].width = 30
        sheet['A1'].font = Font(bold=True)
        sheet['A1'] = 'Date et heure'
        sheet.column_dimensions['B'].width = 20
        sheet['B1'].font = Font(bold=True)
        sheet['B1'] = 'Valeur'
        if is_tht:
            sheet.column_dimensions['C'].width = 20
            sheet['C1'].font = Font(bold=True)
            sheet['C1'] = 'Calcul en mg/m3(n)'

        # add history values to sheet
        for idx, (date_str, value) in enumerate(reversed(tag_history)):
            sheet.cell(row=idx + 2, column=1, value=datetime.fromtimestamp(date_str))
            sheet.cell(row=idx + 2, column=2, value=round(value, 2))
            if is_tht:
                sheet.cell(row=idx + 2, column=3, value=round(value * 3.93, 2))

        # apply alignment to all cells in this sheet
        for column in sheet.columns:
            for cell in column:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # save all to xlsx file
    wb.save(xlsx_file)
    # log end of build 
    logging.info('build done')


# schedule jobs
def export_xlsx_job():
    """Periodic convert of influx data to workbook"""
    # log job run
    logging.info('start export_xlsx_job')

    # build workbook for last 7 days
    for x in range(1, 8):
        exp_day = date.today() - timedelta(days=x)
        try:
            build_daily_xlsx(xlsx_day=exp_day)
        except Exception as e:
            logging.warning(f'error occur during build: {e}')


if __name__ == '__main__':
    # parse command line args
    parser = argparse.ArgumentParser()
    parser.add_argument('filename', nargs='?', default='', help='excel filename')
    args = parser.parse_args()
    # logging setup
    logging.basicConfig(format='%(asctime)s %(message)s', level=logging.INFO)
    # init influxdb
    ifl_cli = InfluxDBClient(host='localhost', port=8086, database='mydb')
    # immediate produce of xlsx file or start redis periodic export
    if args.filename:
        logging.info('build a workbook with today data')
        build_daily_xlsx(xlsx_day=date.today(), to_file=args.filename)
    else:
        logging.info('xlsx-build started')
        # schedule setup
        schedule.every(1).day.at('00:00:30').do(export_xlsx_job)
        # main loop
        while True:
            schedule.run_pending()
            time.sleep(1.0)
