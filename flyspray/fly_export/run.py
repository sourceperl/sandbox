#!/usr/bin/env python3

"""
Export data from flyspray to a CSV file.

"""

import argparse
from codecs import BOM_UTF8
import csv
from datetime import datetime, timedelta, timezone
import logging
from os.path import join
import time
from zoneinfo import ZoneInfo
# apt install python3-openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
# apt install python3-pymysql
import pymysql
# apt install python3-schedule
import schedule
from private_data import DB_USER, DB_PWD, PUB_PATH


# some const
SCHED_H = '00:45'
BUILD_CSV = True
BUILD_XLSX = True


# some const
SQL = """
SELECT  t.task_id, p.project_title, c.category_name, s.status_name, tt.tasktype_name,
        t.item_summary, t.detailed_desc,
        t.date_opened, t.date_closed, t.last_edited_time,
        uo.user_name AS opened_by_user,
        ua.user_name AS assigned_user,
        COUNT(cm.comment_id) AS comments_nb
  FROM  flyspray_tasks AS t
  LEFT JOIN  flyspray_projects AS p ON t.project_id = p.project_id
  LEFT JOIN  flyspray_list_category AS c ON t.product_category = c.category_id
  LEFT JOIN  flyspray_list_status AS s ON t.item_status = s.status_id
  LEFT JOIN  flyspray_list_tasktype AS tt ON t.task_type = tt.tasktype_id
  LEFT JOIN  flyspray_users AS uo ON t.opened_by = uo.user_id
  LEFT JOIN  flyspray_assigned AS a ON t.task_id = a.task_id
  LEFT JOIN  flyspray_users AS ua ON a.user_id = ua.user_id
  LEFT JOIN  flyspray_comments AS cm ON t.task_id = cm.task_id
  WHERE t.date_opened BETWEEN {from_ts} and {to_ts} OR t.last_edited_time BETWEEN {from_ts} and {to_ts}
  GROUP BY t.task_id
  ORDER BY t.task_id DESC LIMIT 5000
"""

SQL_FIELDS = ['task_id', 'project_title', 'category_name', 'status_name', 'tasktype_name', 'opened_by_user', 'last_edited_time', 'date_opened', 'date_closed',
              'assigned_user', 'comments_nb', 'item_summary', 'detailed_desc']


# some class
class ExcelFr(csv.excel):
    delimiter = ';'


csv.register_dialect('excel-fr', ExcelFr())


# some function
def db2csv(db: str, fly_id: str, year: int):
    # define export params
    path_file_csv = join(PUB_PATH, f'annual_csv/fly_{fly_id}_{year}.csv')
    path_file_xlsx = join(PUB_PATH, f'annual_xlsx/fly_{fly_id}_{year}.xlsx')
    sql_open_from_ts = round(datetime(year, 1, 1, 0, 0, 0).timestamp())
    sql_open_to_ts = round(datetime(year, 12, 31, 23, 59, 59).timestamp())

    # connect to DB
    db = pymysql.connect(db=db,
                         user=DB_USER,
                         password=DB_PWD,
                         charset='utf8mb4',
                         cursorclass=pymysql.cursors.DictCursor)

    with db.cursor() as cursor:
        # do request
        sql_req = SQL.format(from_ts=sql_open_from_ts, to_ts=sql_open_to_ts)
        if cursor.execute(sql_req):
            # store sql result as a list of dicts
            sql_row_l = cursor.fetchall()
            # to csv file
            if BUILD_CSV:
                with open(path_file_csv, 'w') as csv_file:
                    # write an UTF-8 BOM (for overide default charset on Excel)
                    csv_file.write(BOM_UTF8.decode('utf8'))
                    # build CSV
                    w = csv.DictWriter(csv_file, fieldnames=SQL_FIELDS,
                                       extrasaction='ignore', dialect='excel-fr')
                    # add header line
                    w.writeheader()
                    # add lines
                    for row_d in sql_row_l:
                        # copy data to edit it
                        csv_row_d = row_d.copy()
                        # update datetime fields (timestamp -> str date)
                        for (k, v) in csv_row_d.items():
                            if k.startswith('date_') or k.endswith('_time'):
                                ts = int(v)
                                if ts > 0:
                                    csv_row_d[k] = datetime.utcfromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
                                else:
                                    csv_row_d[k] = ''
                        # add current line to CSV
                        w.writerow(csv_row_d)
            # to xlsx file
            if BUILD_XLSX:
                # init workbook
                wb = Workbook()
                # init the default first sheet
                ws1 = wb.active
                ws1.title = 'Feuil1'
                col_idx = 0
                # column 1
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 15.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Task ID'
                # column 2
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 15.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Equipe DTS'
                # column 3
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 30.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Date et heure d\'ouverture'
                # column 4
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 30.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Date et heure de fermeture'
                # column 5
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 75.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Résumé'
                # column 6
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 25.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Etat du ticket'
                # column 7
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 30.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Equipe d\'exploitation'
                # column 8
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 30.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Date et heure dernier événement'
                # column 9
                col_idx += 1
                ws1.column_dimensions[get_column_letter(col_idx)].width = 10.0
                ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws1.cell(row=1, column=col_idx).value = 'Nb événement'
                # add rows
                for row_idx, row_d in enumerate(sql_row_l):
                    row_idx += 2
                    col_idx = 0
                    # column 1
                    col_idx += 1
                    task_id = row_d['task_id']
                    ws1.cell(row=row_idx, column=col_idx).style = 'Hyperlink'
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    ws1.cell(row=row_idx, column=col_idx).hyperlink = f'https://flyspray.cloudgaz.net/{fly_id}/{task_id}'
                    ws1.cell(row=row_idx, column=col_idx).value = task_id
                    # column 2
                    col_idx += 1
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    ws1.cell(row=row_idx, column=col_idx).value = row_d['project_title']
                    # column 3
                    col_idx += 1
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    if row_d['date_opened']:
                        opened_dt = datetime.fromtimestamp(row_d['date_opened'], tz=timezone.utc)
                        opened_dt = opened_dt.astimezone(ZoneInfo('Europe/Paris')).replace(tzinfo=None)
                        ws1.cell(row=row_idx, column=col_idx).value = opened_dt
                        ws1.cell(row=row_idx, column=col_idx).number_format = 'dd/mm/yyyy hh:mm'
                    # column 4
                    col_idx += 1
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    if row_d['date_closed']:
                        date_closed = datetime.fromtimestamp(row_d['date_closed'], tz=timezone.utc)
                        date_closed = date_closed.astimezone(ZoneInfo('Europe/Paris')).replace(tzinfo=None)
                        ws1.cell(row=row_idx, column=col_idx).value = date_closed
                        ws1.cell(row=row_idx, column=col_idx).number_format = 'dd/mm/yyyy hh:mm'
                    # column 5
                    col_idx += 1
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    ws1.cell(row=row_idx, column=col_idx).value = row_d['item_summary']
                    # column 6
                    col_idx += 1
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    ws1.cell(row=row_idx, column=col_idx).value = row_d['status_name']
                    # column 7
                    col_idx += 1
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    ws1.cell(row=row_idx, column=col_idx).value = row_d['category_name']
                    # column 8
                    col_idx += 1
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    if row_d['last_edited_time']:
                        last_edited_time = datetime.fromtimestamp(row_d['last_edited_time'], tz=timezone.utc)
                        last_edited_time = last_edited_time.astimezone(ZoneInfo('Europe/Paris')).replace(tzinfo=None)
                        ws1.cell(row=row_idx, column=col_idx).value = last_edited_time
                        ws1.cell(row=row_idx, column=col_idx).number_format = 'dd/mm/yyyy hh:mm'
                    # column 9
                    col_idx += 1
                    ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                    ws1.cell(row=row_idx, column=col_idx).value = row_d['comments_nb']
                # add a default style
                tab = Table(displayName='Table1', ref=f'A1:{get_column_letter(ws1.max_column)}{ws1.max_row}')
                tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium14', showFirstColumn=False,
                                                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws1.add_table(tab)
                # save all to xlsx file
                wb.save(filename=path_file_xlsx)


# schedule job(s)
def export_job():
    # csv year export target = yesterday year
    exp_year = (datetime.now() - timedelta(days=1)).year
    # logging job startup
    logging.info(f'start of export job for year {exp_year}')
    # export all databases for this target year
    for db_name, fly_id in [('flyspray-tca', 'tca'), ('flyspray-tne', 'tne'),
                            ('flyspray-trm', 'trm'), ('flyspray-tvs', 'tvs')]:
        try:
            db2csv(db=db_name, fly_id=fly_id, year=exp_year)
        except Exception as e:
            logging.error(f'error occur during csv build: {e}')
    # logging job end
    logging.info(f'end of export job')


if __name__ == '__main__':
    # parse command line args
    parser = argparse.ArgumentParser()
    parser.add_argument('-r', '--rebuild', action='store_true', help='rebuild mode (from 2010 to now, then exit)')
    args = parser.parse_args()
    # logging setup
    logging.basicConfig(format='%(asctime)s %(message)s', level=logging.INFO)
    # avoid "schedule" module pollute the main log
    logging.getLogger('schedule').propagate = False

    # for rebuild purpose
    if args.rebuild:
        logging.info('start rebuild')
        for y in range(2010, datetime.now().year + 1):
            logging.info(f'rebuild year {y}')
            db2csv(db='flyspray-tca', fly_id='tca', year=y)
            db2csv(db='flyspray-tne', fly_id='tne', year=y)
            db2csv(db='flyspray-trm', fly_id='trm', year=y)
            db2csv(db='flyspray-tvs', fly_id='tvs', year=y)
        logging.info('end of rebuild -> exit')
        exit()

    # daemon mode start msg
    logging.info('start in daemon mode')

    # init schedule (WARN: server local time is UTC time)
    logging.info(f'export job schedule every day at {SCHED_H}')
    schedule.every().day.at(SCHED_H).do(export_job)

    # main loop
    while True:
        schedule.run_pending()
        time.sleep(1)

