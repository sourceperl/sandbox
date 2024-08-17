#!/usr/bin/env python3

"""
Export data from flyspray to xlsx dashboards.

"""

import argparse
from dataclasses import dataclass
from datetime import datetime, timezone
import logging
from os.path import join
import time
from typing import List
from zoneinfo import ZoneInfo
# apt install python3-openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
# apt install python3-pymysql
import pymysql
# apt install python3-schedule
import schedule
from private_data import DB_USER, DB_PWD, PUB_PATH


# some const
SCH_DHB_OPEN_TASK_EVERY_MN = 15


# some class
@dataclass
class TaskInfo:
    fly_id: str
    task_id: int
    project: str
    summary: str
    status: str
    expl_team: str
    evt_nb: int
    open_dt: datetime = None
    last_evt_dt: datetime = None


# some function
def build_dsh_open_tasks():
    SQL = """
    SELECT  t.task_id, p.project_title, c.category_name, s.status_name, tt.tasktype_name,
            t.item_summary, t.detailed_desc,
            t.date_opened, t.date_closed, t.last_edited_time,
            uo.user_name AS opened_by_user,
            ua.user_name AS assigned_user,
            COUNT(cm.comment_id) AS comments_nb
    FROM  flyspray_tasks AS t
    LEFT JOIN  flyspray_projects AS p ON t.project_id = p.project_id
    LEFT JOIN  flyspray_list_category AS c ON t.product_category = c.category_id
    LEFT JOIN  flyspray_list_status AS s ON t.item_status = s.status_id
    LEFT JOIN  flyspray_list_tasktype AS tt ON t.task_type = tt.tasktype_id
    LEFT JOIN  flyspray_users AS uo ON t.opened_by = uo.user_id
    LEFT JOIN  flyspray_assigned AS a ON t.task_id = a.task_id
    LEFT JOIN  flyspray_users AS ua ON a.user_id = ua.user_id
    LEFT JOIN  flyspray_comments AS cm ON t.task_id = cm.task_id
    WHERE t.is_closed = 0
    GROUP BY t.task_id
    ORDER BY t.task_id DESC LIMIT 10000
    """
    # init list of TaskItem
    open_task_l: List[TaskInfo] = []

    for db_name, fly_id in [('flyspray-tca', 'tca'), ('flyspray-tne', 'tne'),
                            ('flyspray-trm', 'trm'), ('flyspray-tvs', 'tvs')]:
        # connect to DB
        db = pymysql.connect(db=db_name,
                             user=DB_USER,
                             password=DB_PWD,
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)
        # SQL request
        with db.cursor() as cursor:
            if cursor.execute(SQL):
                # process sql result
                for sql_d in cursor.fetchall():
                    # init TaskInfo
                    task_info = TaskInfo(fly_id=fly_id,
                                         task_id=sql_d['task_id'],
                                         project=sql_d['project_title'],
                                         summary=sql_d['item_summary'],
                                         status=sql_d['status_name'],
                                         expl_team=sql_d['category_name'],
                                         evt_nb=sql_d['comments_nb'],
                                         )
                    # add localized datetime items if defined
                    if sql_d['date_opened']:
                        as_utc_dt = datetime.fromtimestamp(sql_d['date_opened'], tz=timezone.utc)
                        as_local_dt = as_utc_dt.astimezone(ZoneInfo('Europe/Paris')).replace(tzinfo=None)
                        task_info.open_dt = as_local_dt
                    if sql_d['last_edited_time']:
                        as_utc_dt = datetime.fromtimestamp(sql_d['last_edited_time'], tz=timezone.utc)
                        as_local_dt = as_utc_dt.astimezone(ZoneInfo('Europe/Paris')).replace(tzinfo=None)
                        task_info.last_evt_dt = as_local_dt
                    # populate task_l
                    open_task_l.append(task_info)
        # sort task_l by last event dt
        open_task_l = sorted(open_task_l, key=lambda x: x.last_evt_dt, reverse=True)

    # build xlsx if task_l is not empty
    if open_task_l:
        # init workbook
        wb = Workbook()
        # init the default first sheet
        sheet = wb.active
        sheet.title = 'Feuil1'
        # columns width
        sheet.column_dimensions['A'].width = 15.0
        sheet.column_dimensions['B'].width = 15.0
        sheet.column_dimensions['C'].width = 20.0
        sheet.column_dimensions['D'].width = 30.0
        sheet.column_dimensions['E'].width = 30.0
        sheet.column_dimensions['F'].width = 75.0
        sheet.column_dimensions['G'].width = 25.0
        sheet.column_dimensions['H'].width = 30.0
        sheet.column_dimensions['I'].width = 15.0
        # columns labels
        sheet['A1'].value = 'Instance'
        sheet['B1'].value = 'Task ID'
        sheet['C1'].value = 'Equipe DTS'
        sheet['D1'].value = 'Date et heure d\'ouverture'
        sheet['E1'].value = 'Date et heure dernier événement'
        sheet['F1'].value = 'Résumé'
        sheet['G1'].value = 'Etat du ticket'
        sheet['H1'].value = 'Equipe d\'exploitation'
        sheet['I1'].value = 'Nb événement'
        # add rows
        for row_idx, task_info in enumerate(open_task_l):
            # column A
            cell = sheet.cell(column=column_index_from_string('A'), row=row_idx+2)
            cell.value = task_info.fly_id.upper()
            # column B
            cell = sheet.cell(column=column_index_from_string('B'), row=row_idx+2)
            cell.style = 'Hyperlink'
            cell.hyperlink = f"https://flyspray.cloudgaz.net/{task_info.fly_id}/{task_info.task_id}"
            cell.value = task_info.task_id
            # column C
            cell = sheet.cell(column=column_index_from_string('C'), row=row_idx+2)
            cell.value = task_info.project
            # column D
            cell = sheet.cell(column=column_index_from_string('D'), row=row_idx+2)
            if task_info.open_dt:
                cell.value = task_info.open_dt
                cell.number_format = 'dd/mm/yyyy hh:mm'
            # column E
            cell = sheet.cell(column=column_index_from_string('E'), row=row_idx+2)
            if task_info.last_evt_dt:
                cell.value = task_info.last_evt_dt
                cell.number_format = 'dd/mm/yyyy hh:mm'
            # column F
            cell = sheet.cell(column=column_index_from_string('F'), row=row_idx+2)
            cell.value = task_info.summary
            # column G
            cell = sheet.cell(column=column_index_from_string('G'), row=row_idx+2)
            cell.value = task_info.status
            # column H
            cell = sheet.cell(column=column_index_from_string('H'), row=row_idx+2)
            cell.value = task_info.expl_team
            # column I
            cell = sheet.cell(column=column_index_from_string('I'), row=row_idx+2)
            cell.value = f'{task_info.evt_nb}'
        # apply alignment to all cells
        for col in sheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='center')
        # add a default style to sheet
        tab = Table(displayName='Table1', ref=f'A1:{get_column_letter(sheet.max_column)}{sheet.max_row}')
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium14', showFirstColumn=False,
                                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(tab)
        # save all to xlsx file
        wb.save(filename=join(PUB_PATH, f'fly_all_open_tasks.xlsx'))


# schedule job(s)
def dsh_open_tasks_job():
    # logging job startup
    logging.info(f'start of open tasks dashboard job')
    # do
    try:
        build_dsh_open_tasks()
    except Exception as e:
        logging.error(f'error occur during dashboard build: {e}')
    # logging job end
    logging.info(f'end of job')


if __name__ == '__main__':
    # parse command line args
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--single-shot', action='store_true', help='run in single shot mode')
    args = parser.parse_args()
    # logging setup
    logging.basicConfig(format='%(asctime)s %(message)s', level=logging.INFO)
    # avoid "schedule" module pollute the main log
    logging.getLogger('schedule').propagate = False

    # for rebuild purpose
    if args.single_shot:
        logging.info('run in single shot mode')
        build_dsh_open_tasks()
        logging.info('exit')
        exit()

    # daemon mode start msg
    logging.info('start in daemon mode')

    # init schedule (WARN: server local time is UTC time)
    logging.info(f'dashboard open task build job schedule every {SCH_DHB_OPEN_TASK_EVERY_MN} mn')
    schedule.every(SCH_DHB_OPEN_TASK_EVERY_MN).minutes.do(dsh_open_tasks_job)

    # main loop
    while True:
        schedule.run_pending()
        time.sleep(1)

